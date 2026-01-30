# escala_mes_views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from datetime import datetime, date, timedelta
from calendar import monthrange, day_name, month_name
import calendar
from ..models import (
    Matricula, EventoEscala, EscalaMensal, DiaEscala, 
    ControleSemanal, TipoEvento, Hospital, Setor
)

@login_required
def escala_mes_view(request, mes=None, ano=None):
    """
    View principal para exibir escala mensal no formato de tabela
    similar ao Excel fornecido
    """
    # Determinar mês e ano atual se não especificados
    hoje = datetime.now()
    if mes is None:
        mes = hoje.month
    if ano is None:
        ano = hoje.year

    # Converter para inteiros
    mes = int(mes)
    ano = int(ano)

    # Obter dados da escala mensal do banco
    escala = None
    try:
        # Tentar buscar a escala pelo mês e ano. Se houver mais de uma (hospital/setor), o filtro será aplicado depois.
        # Por enquanto, pegamos a primeira disponível ou filtramos por hospital/setor se fornecido
        hospital_id = request.GET.get('hospital')
        setor_id = request.GET.get('setor')
        
        filtro_escala = Q(mes=mes, ano=ano)
        if hospital_id:
            filtro_escala &= Q(hospital_id=hospital_id)
        if setor_id:
            filtro_escala &= Q(setor_id=setor_id)
            
        escala = EscalaMensal.objects.filter(filtro_escala).first()
    except Exception:
        pass

    # Filtrar por hospital/setor do usuário se não for admin
    user = request.user
    perfil = getattr(user, 'cal_perfil', None)

    if user.is_superuser or user.is_staff:
        profissionais = Matricula.objects.filter(ativo=True)
        hospitais = Hospital.objects.all()
        setores = Setor.objects.all()
    elif perfil and perfil.tipo == 'ESCALANTE' and hasattr(user, 'matricula'):
        profissionais = Matricula.objects.filter(
            hospital=user.matricula.hospital,
            setor=user.matricula.setor,
            ativo=True
        )
        hospitais = Hospital.objects.filter(id=user.matricula.hospital.id)
        setores = Setor.objects.filter(id=user.matricula.setor.id)
    else:
        profissionais = Matricula.objects.filter(ativo=True)
        hospitais = Hospital.objects.all()
        setores = Setor.objects.all()

    # Filtrar por hospital/setor se especificado
    hospital_id = request.GET.get('hospital')
    setor_id = request.GET.get('setor')

    if hospital_id:
        profissionais = profissionais.filter(hospital_id=hospital_id)
    if setor_id:
        profissionais = profissionais.filter(setor_id=setor_id)

    # Calcular datas do mês
    primeiro_dia = date(ano, mes, 1)
    ultimo_dia = date(ano, mes, monthrange(ano, mes)[1])

    # Gerar datas para o cabeçalho baseadas nos dias reais do mês
    datas_cabecalho = []
    # Weekday names in Portuguese
    weekdays_pt = ['SEG', 'TER', 'QUA', 'QUI', 'SEX', 'SAB', 'DOM']
    
    for dia in range(1, ultimo_dia.day + 1):
        data_dia = date(ano, mes, dia)
        # weekday() returns 0 for Monday, 6 for Sunday
        dia_semana_idx = data_dia.weekday()
        datas_cabecalho.append({
            'dia': dia,
            'dia_semana': weekdays_pt[dia_semana_idx],
            'data_completa': data_dia,
            'weekday_idx': dia_semana_idx
        })

    # Calcular totais semanais baseados na virada de semana (Domingo)
    # Processar cada profissional
    dados_profissionais = []
    for profissional in profissionais:
        # Obter eventos deste mês para o profissional
        if escala:
            dias_escala = DiaEscala.objects.filter(
                escala=escala,
                profissional=profissional
            ).order_by('data')
        else:
            eventos = EventoEscala.objects.filter(
                profissional=profissional,
                data__range=[primeiro_dia, ultimo_dia]
            ).order_by('data')
            dias_escala = []
            for evento in eventos:
                dias_escala.append({
                    'data': evento.data,
                    'turnos': evento.tipo_evento.codigo if hasattr(evento, 'tipo_evento') else (evento.tipo.codigo if hasattr(evento, 'tipo') else ''),
                    'horas_dia': evento.tipo_evento.horas if hasattr(evento, 'tipo_evento') else (evento.tipo.horas if hasattr(evento, 'tipo') else 0),
                    'e_tpd': False
                })

        dias_dict = {}
        for dia_obj in dias_escala:
            if isinstance(dia_obj, dict):
                data_obj = dia_obj['data']
                turnos = dia_obj.get('turnos', '')
                horas = float(dia_obj.get('horas_dia', 0))
                e_tpd = dia_obj.get('e_tpd', False)
            else:
                data_obj = dia_obj.data
                turnos = dia_obj.turnos
                horas = float(dia_obj.horas_dia)
                e_tpd = dia_obj.e_tpd

            dias_dict[data_obj.day] = {
                'turnos': turnos,
                'horas': horas,
                'e_tpd': e_tpd
            }

        # Calcular totais semanais reais (quebra no domingo)
        semanas_totais = {}
        semana_atual = 1
        current_week_hours = 0
        
        for dia_idx, data_head in enumerate(datas_cabecalho):
            dia = data_head['dia']
            if dia in dias_dict:
                current_week_hours += dias_dict[dia]['horas']
            
            # Se for domingo (6) ou último dia do mês, fecha a semana
            if data_head['weekday_idx'] == 6 or dia_idx == len(datas_cabecalho) - 1:
                semanas_totais[semana_atual] = current_week_hours
                semana_atual += 1
                current_week_hours = 0
        
        # Garantir que temos chaves para o template (até 6 semanas possíveis)
        for i in range(1, 7):
            if i not in semanas_totais:
                semanas_totais[i] = 0

        carga_anterior = calcular_carga_anterior(profissional, mes, ano)
        tpd_total = sum(1 for dia in dias_dict.values() if dia['e_tpd']) * 8
        
        dados_profissionais.append({
            'profissional': profissional,
            'dias': dias_dict,
            'semanas_totais': semanas_totais,
            'carga_anterior': carga_anterior,
            'carga_semanal': profissional.carga_horaria_semanal,
            'tpd_total': tpd_total,
            'tpd_noturno_total': 0,
            'total_mes': sum(dia['horas'] for dia in dias_dict.values()),
        })

    # Contexto para template
    context = {
        'mes_atual': mes,
        'ano_atual': ano,
        'mes_nome': month_name[mes],
        'primeiro_dia': primeiro_dia,
        'ultimo_dia': ultimo_dia,
        'dados_profissionais': dados_profissionais,
        'datas_cabecalho': datas_cabecalho,
        'hospitais': hospitais,
        'setores': setores,
        'hospital_filtro': int(hospital_id) if hospital_id else None,
        'setor_filtro': int(setor_id) if setor_id else None,
        'escala': escala,
        'meses': [
            (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
            (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
            (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
        ],
        'anos': range(ano - 5, ano + 3),
        'num_semanas': range(1, semana_atual),
    }

        # Calcular meses anterior e próximo
    if mes == 1:
        mes_ant, ano_ant = 12, ano - 1
    else:
        mes_ant, ano_ant = mes - 1, ano
    
    if mes == 12:
        mes_prox, ano_prox = 1, ano + 1
    else:
        mes_prox, ano_prox = mes + 1, ano

    # Portuguese Month names
    meses_pt = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }

    context.update({
        'mes_anterior': {'mes': mes_ant, 'ano': ano_ant, 'nome': meses_pt[mes_ant]},
        'mes_proximo': {'mes': mes_prox, 'ano': ano_prox, 'nome': meses_pt[mes_prox]},
        'mes_nome': meses_pt[mes],
    })

    return render(request, 'escala/escala_mes.html', context)

def calcular_carga_anterior(profissional, mes, ano):
    """
    Calcula CG ANT (carga horária do mês anterior)
    """
    if mes == 1:
        mes_anterior = 12
        ano_anterior = ano - 1
    else:
        mes_anterior = mes - 1
        ano_anterior = ano

    try:
        # Buscar escala do mês anterior
        escala_anterior = EscalaMensal.objects.get(
            mes=mes_anterior,
            ano=ano_anterior,
            hospital=profissional.hospital,
            setor=profissional.setor
        )

        # Calcular total de horas do mês anterior
        total_anterior = DiaEscala.objects.filter(
            escala=escala_anterior,
            profissional=profissional
        ).aggregate(total=Sum('horas_dia'))['total'] or 0

        # Se a carga padrão é 40h e ele fez 48h, CG ANT = +8
        # Se fez 32h, CG ANT = -8
        carga_padrao = 40  # Ou usar profissional.carga_horaria_semanal
        carga_anterior = float(total_anterior) - carga_padrao

        return carga_anterior

    except EscalaMensal.DoesNotExist:
        return 0

@login_required
def importar_escala_excel(request):
    """
    View para importar escala do Excel (similar à que você já tem)
    """
    if request.method == 'POST' and request.FILES.get('arquivo_excel'):
        try:
            arquivo = request.FILES['arquivo_excel']
            mes = int(request.POST['mes'])
            ano = int(request.POST['ano'])
            hospital_id = request.POST['hospital']
            setor_id = request.POST['setor']

            hospital = Hospital.objects.get(id=hospital_id)
            setor = Setor.objects.get(id=setor_id)

            # Verificar se já existe
            escala, created = EscalaMensal.objects.get_or_create(
                mes=mes,
                ano=ano,
                hospital=hospital,
                setor=setor,
                defaults={
                    'criado_por': request.user,
                    'arquivo_excel': arquivo
                }
            )

            if not created:
                escala.arquivo_excel = arquivo
                escala.save()

            # Processar arquivo Excel
            processar_excel_escala(escala, arquivo)

            messages.success(request, f"Escala {mes}/{ano} importada com sucesso!")
            return redirect('cal:escala_mensal_mes_ano', mes=mes, ano=ano)

        except Exception as e:
            messages.error(request, f"Erro ao importar: {str(e)}")

    hospitais = Hospital.objects.all()
    setores = Setor.objects.all()

    hoje = datetime.now()
    context = {
        'hospitais': hospitais,
        'setores': setores,
        'meses': [
            (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
            (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
            (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro')
        ],
        'ano_atual': hoje.year,
    }

    return render(request, 'escala/importar_escala.html', context)

def processar_excel_escala(escala, arquivo):
    """
    Processa o arquivo Excel para popular o banco
    """
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(arquivo, data_only=True)

    # Assumindo que a planilha principal é a primeira
    ws = wb.active

    # Limpar dados antigos
    DiaEscala.objects.filter(escala=escala).delete()
    ControleSemanal.objects.filter(escala=escala).delete()

    # Mapeamento de linhas e colunas (ajustar conforme seu Excel)
    # Baseado na sua estrutura HTML

    linha_profissional_inicio = 5  # Ajustar conforme necessário
    col_nome = 1  # Coluna A
    col_matricula = 2  # Coluna B
    col_cg_ant = 3  # Coluna C
    col_cg_sm = 4  # Coluna D
    col_dia_1 = 5  # Coluna E (dia 1)

    linha_atual = linha_profissional_inicio

    while linha_atual <= ws.max_row:
        nome = ws.cell(row=linha_atual, column=col_nome).value
        matricula = ws.cell(row=linha_atual, column=col_matricula).value

        if nome and matricula:
            try:
                profissional = Matricula.objects.get(matricula=str(matricula))

                # Processar cada dia do mês
                for dia in range(1, 32):
                    col_dia = col_dia_1 + (dia - 1)
                    valor_celula = ws.cell(row=linha_atual, column=col_dia).value

                    if valor_celula:
                        data_dia = date(escala.ano, escala.mes, dia)

                        # Salvar dia da escala
                        DiaEscala.objects.create(
                            escala=escala,
                            profissional=profissional,
                            data=data_dia,
                            turnos=str(valor_celula),
                            horas_dia=calcular_horas_turno(str(valor_celula)),
                            e_tpd='TPD' in str(valor_celula).upper(),
                            e_folga=str(valor_celula).strip() == ''
                        )

                linha_atual += 2  # Pula linha de função/observação

            except Matricula.DoesNotExist:
                print(f"Profissional não encontrado: {matricula}")

        linha_atual += 1

    # Após importar, calcular controles semanais
    calcular_controles_semanais(escala)

def calcular_horas_turno(turno_str):
    """
    Calcula horas baseado na sigla do turno
    """
    if not turno_str or str(turno_str).strip() == '':
        return 0

    turno_str = str(turno_str).upper()

    # Mapeamento simplificado
    if '12' in turno_str:
        return 12
    elif '6' in turno_str:
        return 6
    elif 'TPD' in turno_str:
        return 8
    elif 'FOLGA' in turno_str:
        return 0
    else:
        # Tentar extrair números
        import re
        numeros = re.findall(r'\d+', turno_str)
        if numeros:
            return int(numeros[0])
        return 0

def calcular_controles_semanais(escala):
    """
    Calcula controles semanais após importação
    """
    profissionais = Matricula.objects.filter(
        id__in=DiaEscala.objects.filter(escala=escala)
        .values_list('profissional', flat=True).distinct()
    )

    for profissional in profissionais:
        dias_prof = DiaEscala.objects.filter(
            escala=escala, profissional=profissional
        ).order_by('data')

        # Agrupar por semana (semana 1-5)
        semanas = {1: [], 2: [], 3: [], 4: [], 5: []}

        for dia in dias_prof:
            semana_num = ((dia.data.day - 1) // 7) + 1
            if semana_num > 5:
                semana_num = 5
            semanas[semana_num].append(dia)

        # Criar controles semanais
        for semana_num, dias_semana in semanas.items():
            if dias_semana:
                horas_realizadas = sum(float(dia.horas_dia) for dia in dias_semana)
                horas_tpd = sum(1 for dia in dias_semana if dia.e_tpd) * 8

                ControleSemanal.objects.create(
                    escala=escala,
                    profissional=profissional,
                    semana_numero=semana_num,
                    carga_semanal=profissional.carga_horaria_semanal,
                    carga_anterior=calcular_carga_anterior(
                        profissional, escala.mes, escala.ano
                    ),
                    horas_realizadas=horas_realizadas,
                    horas_tpd=horas_tpd,
                    horas_tpd_noturno=0  # Ajustar conforme necessário
                )
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from calendar import monthrange
import io


@login_required
def exportar_escala_pdf(request, mes, ano):
    

    # 1. Tentar capturar a escala ou avisar que não existe
    # Nota: Idealmente, você deve passar o ID da escala ou hospital/setor via GET
    hospital_id = request.GET.get('hospital')
    setor_id = request.GET.get('setor')
    
    filtros = Q(mes=int(mes), ano=int(ano))
    if hospital_id: filtros &= Q(hospital_id=hospital_id)
    if setor_id: filtros &= Q(setor_id=setor_id)

    escala = EscalaMensal.objects.filter(filtros).first()

    if not escala:
        # Se não achar a escala, redireciona com mensagem de erro em vez de dar 404
        from django.contrib import messages
        messages.error(request, f"Não existe escala gerada para {mes}/{ano}.")
        return redirect('cal:escala_mensal_mes_ano', mes=mes, ano=ano)

    # --- Início da Geração do PDF ---
    buffer = io.BytesIO()
    # Usamos landscape (paisagem) porque 31 dias não cabem em pé
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    
    # Cabeçalho do PDF
    titulo = f"Escala Mensal - {escala.hospital.nome} - {escala.setor.nome} ({mes}/{ano})"
    elements.append(Paragraph(titulo, styles['Title']))
    elements.append(Spacer(1, 12))

    # 2. Montar os dados da Tabela (Cabeçalho de dias)
    num_dias = monthrange(int(ano), int(mes))[1]
    header = ['Profissional'] + [str(d) for d in range(1, num_dias + 1)] + ['Total']
    table_data = [header]

    # 3. Buscar profissionais vinculados a esta escala
    profissionais_ids = DiaEscala.objects.filter(escala=escala).values_list('profissional', flat=True).distinct()
    profissionais = Matricula.objects.filter(id__in=profissionais_ids)

    for prof in profissionais:
        linha = [prof.nome[:15]] # Nome curto para caber
        total_horas = 0
        
        # Buscar dias do profissional nesta escala
        dias = DiaEscala.objects.filter(escala=escala, profissional=prof).order_by('data')
        dias_dict = {d.data.day: d.turnos for d in dias}
        horas_dict = {d.data.day: d.horas_dia for d in dias}

        for d in range(1, num_dias + 1):
            turno = dias_dict.get(d, '-')
            linha.append(turno)
            total_horas += float(horas_dict.get(d, 0))
        
        linha.append(f"{int(total_horas)}h")
        table_data.append(linha)

    # 4. Estilização da Tabela para caber na página
    # Fonte 6 ou 7 é necessária para 31 colunas caberem no A4
    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 1),
        ('RIGHTPADDING', (0, 0), (-1, -1), 1),
    ]))
    
    elements.append(t)
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=escala_{mes}_{ano}.pdf'
    return response

@login_required
def toggle_dia_escala(request, escala_id, profissional_id, dia):
    """
    Alternar entre presença/ausência em um dia específico
    """
    if request.method == 'POST':
        escala = get_object_or_404(EscalaMensal, id=escala_id)
        profissional = get_object_or_404(Matricula, id=profissional_id)

        # Determinar data
        data_dia = date(escala.ano, escala.mes, int(dia))

        # Verificar se já existe
        dia_escala, created = DiaEscala.objects.get_or_create(
            escala=escala,
            profissional=profissional,
            data=data_dia,
            defaults={
                'turnos': 'FOLGA',
                'horas_dia': 0,
                'e_tpd': False,
                'e_folga': True
            }
        )

        if not created:
            if dia_escala.e_folga:
                # Marcar como presente com turno padrão
                dia_escala.turnos = 'SM6'
                dia_escala.horas_dia = 6
                dia_escala.e_folga = False
            else:
                # Marcar como folga
                dia_escala.turnos = 'FOLGA'
                dia_escala.horas_dia = 0
                dia_escala.e_folga = True
            dia_escala.save()

        return redirect('escala_mes_view', mes=escala.mes, ano=escala.ano)

    return redirect('escala_mes_view')