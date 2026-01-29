# escala_views.py
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum
from datetime import datetime, date, timedelta
from calendar import monthrange
from decimal import Decimal

from ..models import (
    EscalaMensal, DiaEscala, ControleSemanal, MapeamentoTurno,
    Matricula, Hospital, Setor, TipoEvento, TPD
)

@login_required
def importar_escala(request):
    """Importa escala do Excel"""
    if request.method == 'POST' and request.FILES.get('arquivo_excel'):
        try:
            arquivo = request.FILES['arquivo_excel']
            mes = int(request.POST['mes'])
            ano = int(request.POST['ano'])
            hospital_id = request.POST['hospital']
            setor_id = request.POST['setor']

            hospital = Hospital.objects.get(id=hospital_id)
            setor = Setor.objects.get(id=setor_id)

            # Verifica se escala já existe
            if EscalaMensal.objects.filter(
                mes=mes, ano=ano, hospital=hospital, setor=setor
            ).exists():
                messages.warning(request, "Escala deste mês já existe. Atualizando...")
                escala = EscalaMensal.objects.get(
                    mes=mes, ano=ano, hospital=hospital, setor=setor
                )
            else:
                # Cria nova escala
                escala = EscalaMensal.objects.create(
                    mes=mes, ano=ano, hospital=hospital, setor=setor,
                    criado_por=request.user,
                    arquivo_excel=arquivo
                )

            # Processa o arquivo Excel
            processar_excel(escala, arquivo)

            messages.success(request, f"Escala importada com sucesso!")
            return redirect('cal:escala_detalhes', escala_id=escala.id)

        except Exception as e:
            messages.error(request, f"Erro ao importar: {str(e)}")

    hospitais = Hospital.objects.all()
    setores = Setor.objects.all()

    context = {
        'hospitais': hospitais,
        'setores': setores,
        'meses': EscalaMensal.MES_CHOICES,
        'ano_atual': datetime.now().year,
    }

    return render(request, 'escala/importar.html', context)

def processar_excel(escala, arquivo):
    """Processa o arquivo Excel e salva no banco"""
    try:
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        if 'Table 1' in wb.sheetnames:
            processar_tabela_principal(escala, wb['Table 1'])
        if 'Planilha2' in wb.sheetnames:
            processar_calculos_semanais(escala, wb['Planilha2'])
        calcular_totais_escala(escala)
    except Exception as e:
        raise Exception(f"Erro ao processar Excel: {str(e)}")

def processar_tabela_principal(escala, sheet):
    """Processa a tabela principal do Excel"""
    dias_colunas = {}
    col_atual = 5  # Coluna E
    for dia in range(1, 32):
        if col_atual <= sheet.max_column:
            cabecalho = sheet.cell(row=3, column=col_atual).value
            if cabecalho and str(cabecalho).isdigit():
                dias_colunas[dia] = col_atual
            col_atual += 2

    DiaEscala.objects.filter(escala=escala).delete()

    for row in range(4, sheet.max_row + 1):
        nome = sheet.cell(row=row, column=1).value
        matricula_cel = sheet.cell(row=row, column=5).value

        if nome and matricula_cel and isinstance(nome, str):
            matricula_str = str(matricula_cel).split()[0] if matricula_cel else ""
            if matricula_str:
                try:
                    profissional = Matricula.objects.get(matricula=matricula_str)
                    for dia, col in dias_colunas.items():
                        try:
                            data = date(escala.ano, escala.mes, dia)
                            valor_celula = sheet.cell(row=row, column=col).value
                            turnos = str(valor_celula) if valor_celula else ""
                            if turnos and turnos.strip():
                                horas = calcular_horas_por_turnos(turnos)
                                DiaEscala.objects.create(
                                    escala=escala,
                                    profissional=profissional,
                                    data=data,
                                    turnos=turnos,
                                    horas_dia=horas,
                                    e_tpd='TPD' in turnos.upper(),
                                    e_folga=horas == 0
                                )
                        except (ValueError, IndexError):
                            continue
                except Matricula.DoesNotExist:
                    continue
                except Exception:
                    continue

def calcular_horas_por_turnos(turnos_str):
    """Calcula total de horas baseado nos turnos"""
    mapeamento_horas = {'SM6': 6, 'ST6': 6, 'SN12': 12, 'M6': 6, 'T6': 6, 'N12': 12, 'TPD': 8, 'FOLGA': 0}
    total = 0
    turnos = turnos_str.replace('<br>', ',').split(',')
    for turno in turnos:
        turno_limpo = turno.strip()
        if turno_limpo in mapeamento_horas:
            total += mapeamento_horas[turno_limpo]
        else:
            import re
            numeros = re.findall(r'\d+', turno_limpo)
            if numeros:
                total += int(numeros[0])
    return total

def processar_calculos_semanais(escala, sheet):
    pass

def calcular_totais_escala(escala):
    ControleSemanal.objects.filter(escala=escala).delete()
    prof_ids = DiaEscala.objects.filter(escala=escala).values_list('profissional', flat=True).distinct()
    profissionais = Matricula.objects.filter(id__in=prof_ids)
    for prof in profissionais:
        dias_prof = DiaEscala.objects.filter(escala=escala, profissional=prof).order_by('data')
        if not dias_prof.exists(): continue
        semanas = {}
        for dia in dias_prof:
            semana_num = ((dia.data.day - 1) // 7) + 1
            if semana_num not in semanas:
                semanas[semana_num] = {'dias': [], 'total_horas': 0}
            semanas[semana_num]['dias'].append(dia)
            semanas[semana_num]['total_horas'] += float(dia.horas_dia)
        for semana_num, dados in semanas.items():
            if 1 <= semana_num <= 6:
                carga_anterior = calcular_carga_anterior(prof, escala.mes, escala.ano)
                ControleSemanal.objects.create(
                    escala=escala, profissional=prof, semana_numero=semana_num,
                    carga_semanal=prof.carga_horaria_semanal, carga_anterior=carga_anterior,
                    horas_realizadas=dados['total_horas']
                )

def calcular_carga_anterior(profissional, mes, ano):
    if mes == 1: mes_ant, ano_ant = 12, ano - 1
    else: mes_ant, ano_ant = mes - 1, ano
    try:
        escala_ant = EscalaMensal.objects.get(mes=mes_ant, ano=ano_ant, hospital=profissional.hospital, setor=profissional.setor)
        total_mes_ant = DiaEscala.objects.filter(escala=escala_ant, profissional=profissional).aggregate(total=Sum('horas_dia'))['total'] or 0
        carga_padrao = profissional.carga_horaria_semanal * 4
        return Decimal(float(total_mes_ant) - carga_padrao)
    except EscalaMensal.DoesNotExist:
        return Decimal(0)

@login_required
def lista_escalas(request):
    escalas = EscalaMensal.objects.all().order_by('-ano', '-mes')
    mes = request.GET.get('mes')
    ano = request.GET.get('ano')
    hospital = request.GET.get('hospital')
    setor = request.GET.get('setor')
    if mes: escalas = escalas.filter(mes=mes)
    if ano: escalas = escalas.filter(ano=ano)
    if hospital: escalas = escalas.filter(hospital_id=hospital)
    if setor: escalas = escalas.filter(setor_id=setor)
    context = {
        'escalas': escalas, 'meses': EscalaMensal.MES_CHOICES,
        'hospitais': Hospital.objects.all(), 'setores': Setor.objects.all(),
        'anos': range(2020, datetime.now().year + 2),
    }
    return render(request, 'escala/lista.html', context)

@login_required
def detalhes_escala(request, escala_id):
    escala = get_object_or_404(EscalaMensal, id=escala_id)
    dias = DiaEscala.objects.filter(escala=escala).order_by('data', 'profissional')
    
    # Lista de abreviações dos dias da semana
    dias_semana_abrev = ['DOM', 'SEG', 'TER', 'QUA', 'QUI', 'SEX', 'SAB']
    
    profissionais = {}
    for dia in dias:
        if dia.profissional.id not in profissionais:
            profissionais[dia.profissional.id] = {
                'profissional': dia.profissional, 
                'dias': [], 
                'total_horas': 0,
                'total_tpd_horas': 0
            }
        
        # Simular processamento de turnos para exibição estilizada
        turnos_list = []
        for t in str(dia.turnos).replace(',', ' ').split():
            sigla = t.strip().upper()
            classe = 'turno-' + sigla.lower()
            if 'TPD' in sigla: classe = 'turno-tpd'
            turnos_list.append({'sigla': sigla, 'classe_css': classe})
        
        dia.turnos_list = turnos_list
        profissionais[dia.profissional.id]['dias'].append(dia)
        profissionais[dia.profissional.id]['total_horas'] += float(dia.horas_dia)
        if dia.e_tpd:
            profissionais[dia.profissional.id]['total_tpd_horas'] += float(dia.horas_dia)

    controles = ControleSemanal.objects.filter(escala=escala).order_by('profissional', 'semana_numero')
    
    num_dias = monthrange(escala.ano, escala.mes)[1]
    # Gerar os dias da semana para o cabeçalho
    cabecalho_dias_semana = []
    for d in range(1, num_dias + 1):
        dt = date(escala.ano, escala.mes, d)
        cabecalho_dias_semana.append(dias_semana_abrev[dt.weekday() if dt.weekday() < 6 else 0]) # Ajuste simples se necessário

    context = {
        'escala': escala, 
        'profissionais_list': profissionais.values(), 
        'controles': controles,
        'total_horas': dias.aggregate(total=Sum('horas_dia'))['total'] or 0,
        'total_tpd': dias.filter(e_tpd=True).count(), 
        'dias_mes': num_dias,
        'dias_semana_abrev': cabecalho_dias_semana,
        'current_day': date.today().day if date.today().month == escala.mes else 0
    }
    return render(request, 'escala/detalhes.html', context)

@login_required
def relatorio_semanal(request, escala_id):
    escala = get_object_or_404(EscalaMensal, id=escala_id)
    semana_num = request.GET.get('semana')
    controles = ControleSemanal.objects.filter(escala=escala)
    if semana_num: controles = controles.filter(semana_numero=semana_num)
    controles = controles.order_by('semana_numero', 'profissional')
    context = {
        'escala': escala, 'controles': controles, 'semanas': range(1, 7),
        'total_horas_realizadas': controles.aggregate(total=Sum('horas_realizadas'))['total'] or 0,
        'total_carga_semanal': controles.aggregate(total=Sum('carga_semanal'))['total'] or 0,
    }
    return render(request, 'escala/relatorio_semanal.html', context)

@login_required
def exportar_escala(request, escala_id):
    escala = get_object_or_404(EscalaMensal, id=escala_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Escala {escala.get_mes_display()}"
    ws['A1'] = f"Escala {escala.get_mes_display()}/{escala.ano}"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=escala_{escala.mes}_{escala.ano}.xlsx'
    wb.save(response)
    return response

@login_required
def dashboard_escala(request):
    ultimas_escalas = EscalaMensal.objects.all().order_by('-ano', '-mes')[:5]
    context = {
        'ultimas_escalas': ultimas_escalas,
        'total_escalas': EscalaMensal.objects.count(),
        'total_profissionais': Matricula.objects.filter(ativo=True).count(),
    }
    return render(request, 'escala/dashboard.html', context)

@login_required
def api_saldo_semanal(request, profissional_id, mes, ano):
    try:
        profissional = Matricula.objects.get(id=profissional_id)
        return JsonResponse({'profissional': profissional.nome_completo, 'saldo': 'disponível'})
    except Matricula.DoesNotExist:
        return JsonResponse({'error': 'Profissional não encontrado'}, status=404)
